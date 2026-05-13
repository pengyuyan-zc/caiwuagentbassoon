# -*- coding: utf-8 -*-
from __future__ import annotations

"""
开票税务凭证生成 - Skill 执行入口（deep-agent 工具）

供 deep-agent 的 LLM 通过工具调用此 skill。

工具列表：
  - invoke_invoice_voucher: 生成开票税务凭证
"""

from pathlib import Path
import json
from typing import Any
from datetime import datetime

import sys as _sys

_SKILL_DIR = Path(__file__).parent
_PROJECT_ROOT = _SKILL_DIR.parent.parent

if str(_PROJECT_ROOT) not in _sys.path:
    _sys.path.insert(0, str(_PROJECT_ROOT))
if str(_SKILL_DIR) not in _sys.path:
    _sys.path.insert(0, str(_SKILL_DIR))

# ── 导入结构化错误类 ──────────────────────────────────────────────────────
from agent.errors import (
    FileFormatError,
    FileNotFoundError,
    SkillExecutionError,
    MissingParameterError,
    error_response,
)


# ─────────────────────────────────────────────────────────────────────
# 业务函数（实际执行逻辑）
# ─────────────────────────────────────────────────────────────────────

def _run_invoice_voucher(
    input_files: list[str] | str,
    output_dir: str | Path | None = None,
) -> dict[str, Any]:
    """
    生成开票税务凭证的内部实现（支持多文件）。

    Args:
        input_files: 开票主体 Excel 文件路径列表（.xlsx 或 .xls），或单个路径字符串。
        output_dir: 输出目录，默认同输入文件。

    Returns:
        dict，包含 success / output_file / voucher_count / line_count / message

    Raises:
        FileNotFoundError: 输入文件不存在
        FileFormatError: 文件格式不正确（不是 Excel）
        SkillExecutionError: 执行过程中出错
    """
    from 开票主体税务凭证生成 import generate_invoice_voucher
    import shutil, tempfile

    # 统一转为列表
    if isinstance(input_files, (str, Path)):
        input_files = [str(input_files)]

    # 验证文件
    for fp in input_files:
        p = Path(fp)
        if not p.exists():
            return error_response(FileNotFoundError(fp))
        header = p.read_bytes()[:8]
        if not header.startswith(b"PK"):
            try:
                text = p.read_text(encoding="utf-8", errors="replace")
                if text.strip().startswith("<"):
                    kind = "HTML"
                else:
                    kind = f"纯文本({len(text.splitlines())}行)"
            except Exception:
                kind = "未知"
            return error_response(FileFormatError(
                fp,
                expected_format="Excel (.xlsx)",
                actual_format=kind,
            ))

    out_dir = Path(output_dir) if output_dir else Path(input_files[0]).parent
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_out_path = out_dir / f"凭证输出_{ts}.xlsx"

    total_vouchers = 0
    total_lines = 0

    if len(input_files) == 1:
        # 单文件直接处理
        result = generate_invoice_voucher(Path(input_files[0]), out_dir)
        return result

    # 多文件：逐个处理后合并输出
    tmp_dirs = []
    tmp_files = []
    for fp in input_files:
        import tempfile
        tmpdir = tempfile.mkdtemp(prefix="voucher_")
        result = generate_invoice_voucher(Path(fp), Path(tmpdir))
        if not result.get("success"):
            return result
        tmp_dirs.append(Path(tmpdir))
        total_vouchers += result.get("voucher_count", 0)
        total_lines += result.get("line_count", 0)
        # 从 tmpdir 找到生成的 xlsx 文件
        xlsx_files = list(Path(tmpdir).glob("凭证输出_*.xlsx"))
        if xlsx_files:
            tmp_files.append(xlsx_files[0])

    # 合并所有凭证文件
    _merge_xlsx_files(tmp_files, final_out_path)

    # 清理临时目录
    import shutil as _shutil
    for td in tmp_dirs:
        try:
            _shutil.rmtree(td)
        except Exception:
            pass

    return {
        "success": True,
        "output_file": str(final_out_path),
        "voucher_count": total_vouchers,
        "line_count": total_lines,
        "message": f"生成完成！共 {len(input_files)} 个文件，{total_vouchers} 张凭证，{total_lines} 条分录"
    }


def _merge_xlsx_files(xlsx_paths: list[Path], output_path: Path) -> None:
    """将多个 xlsx 文件的分录合并到输出文件（以第一个文件为模板）。"""
    import openpyxl
    from openpyxl import load_workbook
    from copy import copy

    wb_out = load_workbook(xlsx_paths[0])
    ws_out = wb_out.active

    # 找到现有数据的最后一行（模板通常有表头行）
    max_row = ws_out.max_row

    for xf in xlsx_paths[1:]:
        wb_src = load_workbook(xf, data_only=True)
        ws_src = wb_src.active
        for row in ws_src.iter_rows(min_row=2, values_only=False):
            new_row = []
            for cell in row:
                new_cell = ws_out.cell(row=max_row + 1, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
            max_row += 1

    ws_out.delete_rows(max_row + 1, amount=ws_out.max_row - max_row)  # 清理空行
    wb_out.save(output_path)


# ─────────────────────────────────────────────────────────────────────
# Deep-agent 工具定义（供 LLM 调用）
# ─────────────────────────────────────────────────────────────────────

def invoke_invoice_voucher(
    input_files: list[str] | str | None = None,
    input_file: str | None = None,
    output_dir: str | None = None,
) -> str:
    """
    生成开票税务凭证（支持多个文件）。根据用户上传的Excel文件（金蝶云格式）自动生成税务凭证。

    参数：
        input_files: 开票主体 Excel 文件路径列表（.xlsx 或 .xls），或单个路径字符串。必须提供。
        output_dir: 凭证输出目录，默认为输入文件所在目录。

    返回：
        执行结果的 JSON 字符串，包含 success、output_file、voucher_count、line_count、message。
        成功示例：{"success": true, "output_file": "...", "voucher_count": 15, "line_count": 45, "message": "生成完成！"}
        失败示例：{"success": false, "error": {"code": "FILE_FORMAT_ERROR", "message": "..."}}
    """
    # 兼容 input_file（单数）和 input_files（复数）
    effective_files: list[str] | str | None = input_files
    if not effective_files and input_file:
        effective_files = input_file
    if not effective_files:
        return json.dumps(error_response(MissingParameterError(
            "invoke_invoice_voucher",
            "input_files",
        )), ensure_ascii=False)
    try:
        for attempt in range(2):
            result = _run_invoice_voucher(
                input_files=effective_files,
                output_dir=output_dir,
            )
            if result.get("success"):
                return json.dumps(result, ensure_ascii=False)
            # 检查是否是可重试的错误
            err_data = result.get("error", {})
            if err_data.get("code") not in ("FILE_NOT_FOUND", "FILE_FORMAT_ERROR", "MISSING_PARAMETER"):
                # 其他错误可能可以重试
                if attempt == 0:
                    import time
                    time.sleep(2)
                continue
            return json.dumps(result, ensure_ascii=False)
        return json.dumps(result, ensure_ascii=False)
    except Exception as e:
        import traceback
        return json.dumps(error_response(SkillExecutionError(
            "invoke_invoice_voucher",
            reason=str(e),
            traceback=traceback.format_exc(),
        )), ensure_ascii=False)


# ─────────────────────────────────────────────────────────────────────
# 工具元数据（deep-agent 读取并注册到 LLM）
# ─────────────────────────────────────────────────────────────────────

TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "invoke_invoice_voucher",
            "description": (
                "生成开票税务凭证（支持多个文件）。根据用户上传的Excel文件（金蝶云格式）自动生成税务凭证。"
                "凭证字=记，账簿按模板默认值，生成借贷分录：借 1122 应收账款 / 贷 6001 主营业务收入 / 贷 2221.01.02 销项税额。"
                "如果用户上传了多个文件，会依次处理并合并输出。"
                "触发词：生成开票税务凭证、生成税务凭证、X月开票税务凭证、开票主体税务凭证、发票税务凭证。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "input_files": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "开票主体Excel文件的完整路径列表（.xlsx 或 .xls），必填。示例：[\"C:/data/1月开票.xlsx\", \"C:/data/2月开票.xlsx\"]"
                    },
                    "output_dir": {
                        "type": "string",
                        "description": "凭证输出目录，默认为输入文件所在目录"
                    }
                },
                "required": ["input_files"]
            }
        }
    }
]
