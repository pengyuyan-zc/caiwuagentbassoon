# -*- coding: utf-8 -*-
"""管理帐开票凭证生成 - 按利润中心分类 + 业务类型细分"""
import argparse
import os
import openpyxl
import re
import json
from collections import defaultdict
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

parser = argparse.ArgumentParser()
parser.add_argument('--输入', type=str, default=None)
parser.add_argument('--输出', type=str, default=None)
parser.add_argument('--客户', type=str, default=None)
parser.add_argument('--模板', type=str, default=None)
args = parser.parse_args()

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PARENT_DIR = os.path.dirname(_SCRIPT_DIR)  # 管理帐 目录

OA_FILE   = args.输入 if args.输入 else os.path.join(_PARENT_DIR, '1112月部分开票.xlsx')
OUT_FILE  = args.输出 if args.输出 else os.path.join(_PARENT_DIR, '管理帐凭证_测试.xlsx')
CUST_FILE = args.客户 if args.客户 else os.path.join(_PARENT_DIR, '客户明细管理帐.xlsx')
TPL_FILE  = args.模板 if args.模板 else os.path.join(_PARENT_DIR, '202510月凭证-管理.xlsx')

# ========================
# 列索引（0-based）
# ========================
COL_SUMM   = 20
COL_ACCT_C = 21
COL_ACCT_N = 22
COL_CCY_C  = 69
COL_ORIG   = 78
COL_DEBIT  = 79
COL_CREDIT = 80
COL_CUST_C = 59
COL_CUST_N = 60
COL_DATE   = 3
COL_YEAR   = 5
COL_PERIOD = 6
COL_VWORD  = 7
COL_VNO    = 9
COL_BOOK_C = 1
COL_SEQ    = 0

# ========================
# 1. 客户明细
# ========================
print('1. 加载客户明细...')
wb_c = openpyxl.load_workbook(CUST_FILE, data_only=True)
ws_c = wb_c[wb_c.sheetnames[0]]
cust_full = {}
for r in range(2, ws_c.max_row + 1):
    code = str(ws_c.cell(r, 1).value or '').strip()
    name = str(ws_c.cell(r, 2).value or '').strip()
    if code and name: cust_full[name] = code

def core_name(name):
    name = str(name).strip()
    n = re.sub(r'[(（].*?[)）]', '', name).strip()
    n = re.sub(r'^(广州|深圳|佛山|中山|珠海|东莞|惠州|江门|肇庆|湛江|茂名|韶关|梅州|汕头|汕尾|河源|阳江|清远|潮州|揭阳|云浮|广东|广西|湖南|湖北|海南|江西|浙江|江苏|安徽|福建|山东|河南|河北|四川|重庆|北京|上海|天津|云南|贵州|陕西|山西|内蒙|新疆|西藏|宁夏|青海|甘肃|黑龙江|吉林|辽宁)(市|省)?', '', n)
    n = re.sub(r'(有限公司|股份有限公司|有限责任公司|分公司|子公司|集团股份有限公司)$', '', n).strip()
    return n

cust_core = {}
for name, code in cust_full.items():
    c = core_name(name)
    if c and c not in cust_core: cust_core[c] = code

STOP_WORDS = {
    '有限公司', '股份有限公司', '有限责任公司', '集团', '集团有限公司',
    '分公司', '子公司', '中国', '省', '市', '县', '区',
    '镇', '乡', '村', '路', '号', '街', '道', '广场',
    '大厦', '中心', '园区', '楼', '栋', '层',
    '-', '—', '–', '·', '•', '～',
    '（', '）', '(', ')', '「', '」', '『', '』',
    ' ', '  ', '　',
}

def extract_keywords(name):
    """提取名称的关键词集合（去掉停用词后按标点拆分）"""
    name = str(name).strip()
    for sw in STOP_WORDS:
        name = name.replace(sw, '')
    name = re.sub(r'\d+', '', name)
    parts = re.split(r'[、，。；,.;+/@#%&]+', name)
    keywords = set()
    for part in parts:
        part = part.strip()
        if len(part) >= 2:
            keywords.add(part)
    return keywords


def fuzzy_score(oa_keywords, cust_keywords):
    """计算OA关键词与客户关键词的相似度得分"""
    if not oa_keywords or not cust_keywords:
        return 0.0
    intersection = oa_keywords & cust_keywords
    if not intersection:
        return 0.0
    coverage = len(intersection) / len(oa_keywords)
    bonus = min(len(intersection) - 1, 3) * 0.05 if len(intersection) > 1 else 0
    return coverage + bonus


def fuzzy_match_best(oa_name, oa_full=None, threshold=0.4):
    """模糊关键词匹配（兜底方案），返回最佳客户编码、名称和得分"""
    base_keywords = extract_keywords(oa_name)
    if oa_full:
        base_keywords |= extract_keywords(oa_full)
    if not base_keywords:
        return None

    best_score = 0.0
    best_code = None
    best_match_name = None

    for cust_name, cust_code in cust_full.items():
        cust_keywords = extract_keywords(cust_name)
        score = fuzzy_score(base_keywords, cust_keywords)
        if score > best_score:
            best_score = score
            best_code = cust_code
            best_match_name = cust_name

    if best_score >= threshold:
        return best_code, best_match_name, round(best_score, 2)
    return None


def find_cust(oa_name, oa_full=None):
    if not oa_name: return '未匹配', oa_name
    oa_name = str(oa_name).strip()
    if oa_name in cust_full: return cust_full[oa_name], oa_name
    oa_np = re.sub(r'[(（].*?[)）]', '', oa_name).strip()
    if oa_np in cust_full: return cust_full[oa_np], oa_np
    c = core_name(oa_np)
    if c and c in cust_core: return cust_core[c], oa_name
    if oa_full:
        oa_full = str(oa_full).strip()
        if oa_full in cust_full: return cust_full[oa_full], oa_full
        fnp = re.sub(r'[(（].*?[)）]', '', oa_full).strip()
        if fnp in cust_full: return cust_full[fnp], fnp
        fc = core_name(fnp)
        if fc and fc in cust_core: return cust_core[fc], oa_full
    # 兜底：模糊关键词匹配
    fuzzy_result = fuzzy_match_best(oa_name, oa_full)
    if fuzzy_result:
        code, match_name, score = fuzzy_result
        return code, f'{match_name}[模糊:{score:.0%}]'
    return '未匹配', oa_name

# ========================
# 2. 加载OA
# ========================
print('2. 加载OA...')
wb_o = openpyxl.load_workbook(OA_FILE, data_only=True)
ws_o = wb_o.active
col = {}
for c in range(1, ws_o.max_column + 1):
    v = ws_o.cell(1, c).value
    if v is not None: col[str(v).strip()] = c
rc = lambda n: col.get(n, 0)

def get_date(全额_str, 实际_str):
    for src in [全额_str, 实际_str]:
        if not src or src == 'None': continue
        try:
            for item in json.loads(src):
                for k in ['所属年月', '所属月份']:
                    if k in item:
                        m = re.search(r'(\d{4})[年/\-]?(\d{1,2})', str(item.get(k, '')))
                        if m:
                            y, mo = int(m.group(1)), int(m.group(2))
                            return f'{y}-{mo:02d}-{31 if mo == 12 else 30}', y, mo
        except: pass
    return '2025-10-31', 2025, 10

# ========================
# 3. 解析实际明细
# ========================
def parse_actual_detail(实际明细_str):
    """解析实际明细JSON，返回 [{'项目': 项目名, '金额': float}]"""
    if not 实际明细_str or 实际明细_str in ('None', '[]', ''): return []
    try:
        items = json.loads(str(实际明细_str))
    except: return []
    result = []
    for item in items:
        proj = str(item.get('实际发生项目', '')).strip()
        amt = float(item.get('发生金额（元）', 0) or 0)
        if proj and abs(amt) > 0.01:
            result.append({'项目': proj, '金额': amt})
    return result

# ========================
# 4. 科目模板定义
# ========================
def get_主营业务科目(内部类型):
    if '承揽' in 内部类型: return ('6001.03', '主营业务收入_收入-承揽')
    if '猎头' in 内部类型: return ('6001.04', '主营业务收入_猎头收入（旧）')
    if '灵工' in 内部类型 or '小时间' in 内部类型: return ('6001.10', '主营业务收入_收入-灵工')
    if '代理' in 内部类型: return ('6001.02', '主营业务收入_收入-代理')
    if '派遣' in 内部类型 or '假外包' in 内部类型: return ('6001.01', '主营业务收入_收入-派遣')
    return ('6001.03', '主营业务收入_收入-承揽')

def get_应收款科目(内部类型, is_gd):
    if not is_gd: return ('1122.04', '应收账款_承揽客户应收款（旧）')
    if '猎头' in 内部类型: return ('1122.05', '应收账款_猎头客户应收款（旧）')
    if '灵工' in 内部类型 or '小时间' in 内部类型: return ('1122.11', '应收账款_小时间承揽客户应收款（旧）')
    if '代理' in 内部类型: return ('1122.03', '应收账款_代理客户应收款（旧）')
    if '派遣' in 内部类型 or '假外包' in 内部类型: return ('1122.02', '应收账款_派遣客户应收款（旧）')
    if '承揽' in 内部类型: return ('1122.04', '应收账款_承揽客户应收款（旧）')
    return ('1122.04', '应收账款_承揽客户应收款（旧）')

def get_贷方科目(内部类型, is_gd):
    """非广东用 1221.01.03"""
    if not is_gd: return ('1221.01.03', '其他应收款_内部公司往来_代交易往来款')
    return get_主营业务科目(内部类型)

# ========================
# 5. 实际费用项目 → 科目映射（贷方）
# ========================
def get_credit_acct_from_item(项目名, 内部类型):
    """根据实际费用项目名称返回贷方科目"""
    p = str(项目名)

    # 服务费/顾问费/招聘服务费/不含税金额 → 主营业务收入
    if any(k in p for k in ['服务费', '顾问费', '招聘服务费', '不含税金额', '项目管理', '服务款项']):
        return get_主营业务科目(内部类型)

    # 工伤/养老/医疗/失业/生育/社保/险种
    if any(k in p for k in ['工伤', '养老', '医疗', '失业', '生育', '社保', '险种']):
        return ('2202.02.02', '应付账款_代收代付款项_代付社保')

    # 工资/佣金/派遣服务费(不含税)
    if any(k in p for k in ['工资', '佣金', '不含税']):
        return ('2202.02.01', '应付账款_代收代付款项_代付工资')

    # 公积金/住房
    if any(k in p for k in ['公积金', '住房']):
        return ('2202.02.03', '应付账款_代收代付款项_代付公积金')

    # 商保/商业险 → 主营业务收入
    if any(k in p for k in ['商保', '商业险']):
        return ('6001.01', '主营业务收入_收入-派遣')

    # 个税/个人所得税/个调税
    if any(k in p for k in ['个税', '个人所得税', '个调税']):
        return ('2202.02.05', '应付账款_代收代付款项_代付个税')

    # 残保/残疾人
    if any(k in p for k in ['残保', '残疾人']):
        return ('2202.02.06', '应付账款_代收代付款项_代付残保金')

    # 经营所得
    if '经营所得' in p:
        return ('2202.02.12', '应付账款_代收代付款项_代付经营所得')

    # 代付税金 → 借方 6401.01.08.03（负数），不在贷方
    if any(k in p for k in ['代付税金', '税金']):
        return None

    # 税返/产业园奖励
    if any(k in p for k in ['产业园', '税返', '奖励', '返税']):
        return ('6301.04.01', '营业外收入_税返政策奖励_产业园奖励')

    # 其他默认代付
    return ('2202.02.07', '应付账款_代收代付款项_代付其他款项')


# ========================
# 5. 生成凭证分录
# ========================
def build_entries(凭证, is_gd):
    """生成一条OA对应的所有凭证分录"""
    entries = []
    cust_simple = re.sub(r'[(（].*?[)）]', '', 凭证['客户名称']).strip()
    开票主体 = re.sub(r'[(（].*?[)）]', '', str(凭证.get('开票主体', ''))).strip()
    开票主体 = re.sub(r'^开票主体[：:]\s*', '', 开票主体)
    开票主体 = re.sub(r'^代?开票主体[：:]\s*', '', 开票主体)
    摘要_base = f'{开票主体}开票--{cust_simple}' if 开票主体 else f'开票--{cust_simple}'
    摘要_base = 摘要_base.strip('--')

    总开票额 = 凭证['总开票额']
    内部类型 = 凭证['内部类型']
    实际明细 = parse_actual_detail(凭证.get('实际明细_str', ''))

    if is_gd:
        # ---- 广东: 借 1122 + 贷方按实际明细逐项 ----
        dr_acct = get_应收款科目(内部类型, True)
        entries.append({'摘要': 摘要_base, '科目编码': dr_acct[0], '科目名称': dr_acct[1], '借方': 总开票额, '贷方': 0})

        # 遍历实际明细，每项一个贷方分录；代付税金 → 借方负数
        for item in 实际明细:
            acct = get_credit_acct_from_item(item['项目'], 内部类型)
            if acct is None:
                entries.append({
                    '摘要': 摘要_base,
                    '科目编码': '6401.01.08.03',
                    '科目名称': '主营业务成本_派遣成本_税金费用-派遣_增值税及附加-派遣',
                    '借方': -item['金额'],
                    '贷方': 0,
                })
            else:
                entries.append({'摘要': 摘要_base, '科目编码': acct[0], '科目名称': acct[1], '借方': 0, '贷方': item['金额']})
    else:
        # ---- 非广东: 借 1122 + 贷 1221.01.03 ----
        dr_acct = get_应收款科目(内部类型, False)
        cr_acct = get_贷方科目(内部类型, False)
        entries.append({'摘要': 摘要_base, '科目编码': dr_acct[0], '科目名称': dr_acct[1], '借方': 总开票额, '贷方': 0})
        entries.append({'摘要': 摘要_base, '科目编码': cr_acct[0], '科目名称': cr_acct[1], '借方': 0, '贷方': 总开票额})

    return entries

# ========================
# 7. 处理OA
# ========================
print('3. 处理OA...')
all_voucher_rows = []
seen = set()
matched = unmatched = 0

for r in range(2, ws_o.max_row + 1):
    审批编号 = str(ws_o.cell(r, rc('审批编号')).value or '').strip()
    if not 审批编号 or 审批编号.startswith('~'): continue
    if 审批编号 in seen: continue
    seen.add(审批编号)
    if str(ws_o.cell(r, rc('审批状态')).value or '').strip() != '已完成': continue
    总开票额 = ws_o.cell(r, rc('总开票额')).value
    if not 总开票额 or float(总开票额) == 0: continue

    利润中心 = str(ws_o.cell(r, rc('利润中心')).value or '').strip()
    内部类型 = str(ws_o.cell(r, rc('内部类型分类')).value or '').strip()
    开票主体 = str(ws_o.cell(r, rc('开票主体')).value or '').strip()
    客户名称 = str(ws_o.cell(r, rc('客户/项目名称')).value or '').strip()
    客户全称 = str(ws_o.cell(r, rc('客户公司全称')).value or '').strip()
    实际明细_str = str(ws_o.cell(r, rc('实际明细')).value or '')
    全额明细_str = str(ws_o.cell(r, rc('全额开票部分票面明细')).value or '')

    cust_code, cust_match = find_cust(客户名称, 客户全称)
    总开票额_val = float(总开票额)
    voucher_date, year, period = get_date(全额明细_str, 实际明细_str)

    # 判断广东 vs 非广东
    is_gd = bool(re.search(r'广东|广州', 利润中心))

    if cust_code == '未匹配': unmatched += 1
    else: matched += 1

    all_voucher_rows.append({
        '凭证日期': voucher_date,
        '会计年度': year,
        '期间': period,
        '利润中心': 利润中心,
        '内部类型': 内部类型,
        '开票主体': 开票主体,
        '客户编码': cust_code,
        '客户名称': cust_match,
        '总开票额': 总开票额_val,
        '实际明细_str': 实际明细_str,
        '全额明细_str': 全额明细_str,
        'is_gd': is_gd,
        '未匹配': cust_code == '未匹配',
    })

print(f'   共 {len(all_voucher_rows)} 条, 匹配={matched}, 未匹配={unmatched}')

# 生成所有分录
for vr in all_voucher_rows:
    vr['entries'] = build_entries(vr, vr['is_gd'])

# 统计
stats = defaultdict(lambda: defaultdict(int))
for vr in all_voucher_rows:
    k = '广东' if vr['is_gd'] else '非广东'
    stats[k][vr['内部类型']] += 1

print('\n   统计:')
for k in ['广东', '非广东']:
    print(f'   {k}:')
    for t, c in sorted(stats[k].items()):
        print(f'     {t}: {c}条')

# 预览
print('\n   前10条凭证预览:')
for i, vr in enumerate(all_voucher_rows[:10]):
    print(f'   [{i+1}] {vr["利润中心"]}/{vr["内部类型"]} 客户={vr["客户编码"]}/{vr["客户名称"][:15]} ({len(vr["entries"])}行)')
    for e in vr['entries']:
        dr = f'{e["借方"]:.2f}' if e['借方'] else '-'
        cr = f'{e["贷方"]:.2f}' if e['贷方'] else '-'
        print(f'        [{e["科目编码"]}]{e["科目名称"]} 借={dr} 贷={cr}')

# ========================
# 8. 输出到Excel
# ========================
print('\n4. 输出Excel...')
wb_tpl = openpyxl.load_workbook(TPL_FILE)
ws_tpl = wb_tpl[wb_tpl.sheetnames[0]]

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = ws_tpl.title

for r in range(1, 3):
    for c in range(1, ws_tpl.max_column + 1):
        src = ws_tpl.cell(r, c)
        dst = ws_out.cell(r, c)
        dst.value = src.value
        if src.has_style:
            try:
                dst.font = src.font.copy()
                dst.fill = src.fill.copy()
                dst.border = src.border.copy()
                dst.alignment = src.alignment.copy()
            except: pass

for c in range(1, ws_tpl.max_column + 1):
    col_letter = get_column_letter(c)
    if col_letter in ws_tpl.column_dimensions:
        ws_out.column_dimensions[col_letter].width = ws_tpl.column_dimensions[col_letter].width

RED_FILL = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

total_rows = sum(len(vr['entries']) for vr in all_voucher_rows)
print(f'   共 {len(all_voucher_rows)} 张凭证, {total_rows} 行分录')

row_idx = 2
voucher_no = 0

for vr in all_voucher_rows:
    voucher_no += 1
    entries = vr['entries']
    first_row = True

    for e in entries:
        row_idx += 1

        if first_row:
            ws_out.cell(row_idx, COL_SEQ + 1).value     = voucher_no
            ws_out.cell(row_idx, COL_BOOK_C + 1).value  = '501'
            ws_out.cell(row_idx, COL_DATE + 1).value    = vr['凭证日期']
            ws_out.cell(row_idx, COL_YEAR + 1).value   = vr['会计年度']
            ws_out.cell(row_idx, COL_PERIOD + 1).value  = vr['期间']
            ws_out.cell(row_idx, COL_VWORD + 1).value   = '记'
            ws_out.cell(row_idx, COL_VNO + 1).value    = voucher_no
            first_row = False

        ws_out.cell(row_idx, COL_SUMM + 1).value   = e['摘要']
        ws_out.cell(row_idx, COL_ACCT_C + 1).value = e['科目编码']
        ws_out.cell(row_idx, COL_ACCT_N + 1).value = e['科目名称']
        ws_out.cell(row_idx, COL_CCY_C + 1).value  = 'PRE001'
        ws_out.cell(row_idx, COL_CUST_C + 1).value = vr['客户编码']
        ws_out.cell(row_idx, COL_CUST_N + 1).value = vr['客户名称']

        amt = e['借方'] if e['借方'] else e['贷方']
        ws_out.cell(row_idx, COL_ORIG + 1).value = amt
        if e['借方']: ws_out.cell(row_idx, COL_DEBIT + 1).value  = e['借方']
        if e['贷方']: ws_out.cell(row_idx, COL_CREDIT + 1).value = e['贷方']

        if vr['未匹配']:
            for col_i in range(1, ws_tpl.max_column + 1):
                ws_out.cell(row_idx, col_i).fill = RED_FILL

wb_out.save(OUT_FILE)
print(f'\n保存: {OUT_FILE}')
print(f'共 {voucher_no} 张凭证, {total_rows} 行分录, {ws_tpl.max_column}列')
print('Done!')
