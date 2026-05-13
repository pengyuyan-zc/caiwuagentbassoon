from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

import salary_voucher_generator as base
from salary_voucher_generator import EntryRow


def _get_sheet(wb, preferred_name: str):
    """优先用指定名称的 sheet，找不到则回退到第一个 sheet。"""
    if preferred_name in wb.sheetnames:
        return wb[preferred_name]
    if wb.worksheets:
        return wb.worksheets[0]
    raise KeyError(f"Workbook 中没有任何 sheet，尝试查找: {preferred_name}")


@dataclass(frozen=True)
class ApprovalRow:
    approval_no: str
    customer: str
    profit_center: str
    contract_business_type: str
    internal_business_type: str
    account_book_name: str
    salary_period_start: datetime
    net_salary: Decimal
    tax: Decimal
    approval_record: str
    amount_check: Decimal = Decimal("0")


TARGET_PROFIT_CENTER = "广东公司"
APPROVAL_SHEET_NAME = base.APPROVAL_SHEET_NAME
VOUCHER_SHEET_NAME = base.VOUCHER_SHEET_NAME
ACCOUNT_BOOK_MAPPING_FILE = base.ACCOUNT_BOOK_MAPPING_FILE
CUSTOMER_MAPPING_FILE = "客户帐套_管理.xlsx"
BANK_JOURNAL_FILE = base.BANK_JOURNAL_FILE
ERROR_HEADERS = [
    "审批编号",
    "客户",
    "归属利润中心",
    "内部业务类型",
    "工资所属期起",
    "实发工资",
    "个税",
    "错误原因",
]

ORIGINAL_BUSINESS_ACCOUNT_CODES = {
    "承揽业务": "6401.04.01",
    "派遣业务": "6401.04.01",
    "代理业务": "6401.02.01",
}

PAYMENT_BUSINESS_TYPES = {"派遣业务", "假外包业务", "代理业务"}

ACCOUNT_CODE_NAME_MAP = {
    **base.ACCOUNT_CODE_NAME_MAP,
    "1221.01.03": "其他应收款_内部公司往来_代交易往来款",
    "2202.01.05": r"应付账款_投保申报款项_待分配残疾人\残保金费用",
    "2202.02.01": "应付账款_代收代付款项_代付工资",
    "2202.02.05": "应付账款_代收代付款项_代付个税",
    "2202.02.12": "应付账款_代收代付款项_代付经营所得",
}

# 复用原脚本的凭证行写入函数，该函数从原模块全局表读取科目名称。
base.ACCOUNT_CODE_NAME_MAP.update(ACCOUNT_CODE_NAME_MAP)


def _get_row_value(
    row: tuple[object, ...],
    header_index: dict[str, int],
    header_name: str,
    fallback_index: int,
) -> object:
    index = header_index.get(header_name, fallback_index)
    return row[index] if index < len(row) else None


def _build_original_entry_rows(approval: ApprovalRow) -> list[EntryRow]:
    business_type = approval.internal_business_type
    try:
        expense_code = ORIGINAL_BUSINESS_ACCOUNT_CODES[business_type]
    except KeyError as exc:
        raise ValueError(f"不支持的内部业务类型: {business_type}") from exc

    entries = [
        EntryRow(expense_code, approval.net_salary, Decimal("0"), approval.customer),
        EntryRow("1002", Decimal("0"), approval.net_salary, approval.customer),
    ]
    if approval.tax != Decimal("0"):
        entries.extend(
            [
                EntryRow(expense_code, approval.tax, Decimal("0"), approval.customer, is_tax=True),
                EntryRow("2221.13", Decimal("0"), approval.tax, approval.customer, is_tax=True),
            ]
        )
    return entries


def build_entry_rows(
    approval: ApprovalRow,
    target_profit_center: str = TARGET_PROFIT_CENTER,
) -> list[EntryRow]:
    """将审批记录展开为工资发放管理帐凭证分录。"""
    if approval.profit_center != target_profit_center:
        return [
            EntryRow("1221.01.03", approval.net_salary, Decimal("0"), approval.customer),
            EntryRow("1002", Decimal("0"), approval.net_salary, approval.customer),
        ]

    business_type = approval.internal_business_type

    if business_type == "承揽业务":
        return _build_original_entry_rows(approval)

    if business_type in PAYMENT_BUSINESS_TYPES:
        entries = [
            EntryRow("2202.02.01", approval.net_salary, Decimal("0"), approval.customer),
            EntryRow("1002", Decimal("0"), approval.net_salary, approval.customer),
        ]
        if approval.tax != Decimal("0"):
            entries.extend(
                [
                    EntryRow("2202.02.05", approval.tax, Decimal("0"), approval.customer, is_tax=True),
                    EntryRow("2221.13", Decimal("0"), approval.tax, approval.customer, is_tax=True),
                ]
        )
        return entries

    if business_type == "灵工业务":
        return [
            EntryRow("2202.02.12", approval.net_salary, Decimal("0"), approval.customer),
            EntryRow("1002", Decimal("0"), approval.net_salary, approval.customer),
        ]

    if business_type == "助残业务":
        return [
            EntryRow("2202.01.05", approval.net_salary, Decimal("0"), approval.customer),
            EntryRow("1002", Decimal("0"), approval.net_salary, approval.customer),
        ]

    raise ValueError(f"不支持的内部业务类型: {business_type}")


def load_approval_rows(approval_path: Path) -> tuple[list[ApprovalRow], list[list[object]]]:
    """读取审批表数据。"""
    wb = load_workbook(approval_path, data_only=True)
    ws = _get_sheet(wb, APPROVAL_SHEET_NAME)
    headers = [cell.value for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers) if name}
    approvals: list[ApprovalRow] = []
    errors: list[list[object]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        approval_no_val = _get_row_value(row, header_index, "审批编号", 0)
        if not approval_no_val:
            continue

        try:
            approvals.append(
                ApprovalRow(
                    approval_no=str(approval_no_val).strip(),
                    customer=str(_get_row_value(row, header_index, "客户", 5) or "").strip(),
                    profit_center=str(_get_row_value(row, header_index, "归属利润中心", 3) or "").strip(),
                    contract_business_type=str(_get_row_value(row, header_index, "合同业务类型", 7) or "").strip(),
                    internal_business_type=str(_get_row_value(row, header_index, "内部业务类型", 8) or "").strip(),
                    account_book_name=str(_get_row_value(row, header_index, "发放和申报主体", 9) or "").strip(),
                    salary_period_start=base.parse_datetime_text(_get_row_value(row, header_index, "工资所属期起", 14)),
                    net_salary=base.parse_money(_get_row_value(row, header_index, "实发工资", 24)),
                    tax=base.parse_money(_get_row_value(row, header_index, "个税", 22)),
                    approval_record=str(_get_row_value(row, header_index, "审批记录", 53) or "").strip(),
                    amount_check=base.parse_money(_get_row_value(row, header_index, "金额检验（必须为0）", 25)),
                )
            )
        except ValueError as exc:
            errors.append(
                [
                    approval_no_val,
                    _get_row_value(row, header_index, "客户", 5),
                    _get_row_value(row, header_index, "归属利润中心", 3),
                    _get_row_value(row, header_index, "内部业务类型", 8),
                    _get_row_value(row, header_index, "工资所属期起", 14),
                    _get_row_value(row, header_index, "实发工资", 24),
                    _get_row_value(row, header_index, "个税", 22),
                    str(exc),
                ]
            )
    return approvals, errors


def build_error_row(approval: ApprovalRow, reason: str) -> list[object]:
    """构建错误表行。"""
    return [
        approval.approval_no,
        approval.customer,
        approval.profit_center,
        approval.internal_business_type,
        approval.salary_period_start.strftime("%Y-%m-%d %H:%M"),
        round(float(approval.net_salary), 2),
        round(float(approval.tax), 2),
        reason,
    ]


def write_error_workbook(error_rows: list[list[object]], output_path: Path) -> None:
    """写入错误表。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "错误明细"
    ws.append(ERROR_HEADERS)
    for row in error_rows:
        ws.append(row)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def generate_salary_payment_management_voucher_files(
    approval_path: Path,
    template_path: Path,
    output_dir: Path,
    account_book_mapping_path: Path | None = None,
    customer_mapping_path: Path | None = None,
    target_profit_center: str = TARGET_PROFIT_CENTER,
) -> dict[str, int]:
    """生成工资发放管理帐凭证文件和错误明细。"""
    approvals, preload_errors = load_approval_rows(approval_path)

    account_book_mapping: dict[str, tuple[str, str]] | None = None
    customer_mapping: dict[str, str] | None = None
    bank_mapping = None

    if account_book_mapping_path and account_book_mapping_path.exists():
        account_book_mapping = base.load_account_book_mapping(account_book_mapping_path)
    if customer_mapping_path and customer_mapping_path.exists():
        customer_mapping = base.load_customer_mapping(customer_mapping_path)
    bank_journal_path = Path(BANK_JOURNAL_FILE)
    if bank_journal_path.exists():
        bank_mapping = base.load_bank_account_mapping(bank_journal_path)

    voucher_wb = load_workbook(template_path)
    voucher_ws = voucher_wb[VOUCHER_SHEET_NAME]

    headers = [cell.value for cell in voucher_ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers) if name}

    template_row = [cell.value for cell in voucher_ws[3]]
    base_voucher_seq = int(template_row[header_index.get("FBillHead(GL_VOUCHER)", 0)] or 0)
    base_voucher_no = int(template_row[header_index.get("FVOUCHERGROUPNO", 9)] or 0)
    base_entry_seq = int(template_row[header_index.get("FEntity", 19)] or 0)

    if voucher_ws.max_row > 2:
        voucher_ws.delete_rows(3, voucher_ws.max_row - 2)

    output_rows: list[list[object]] = []
    error_rows = list(preload_errors)
    voucher_offset = 0
    entry_offset = 0

    for approval in approvals:
        if approval.amount_check != Decimal("0"):
            error_rows.append(build_error_row(approval, f"金额检验不为0: {approval.amount_check}"))
            continue

        try:
            entries = build_entry_rows(approval, target_profit_center=target_profit_center)
        except ValueError as exc:
            error_rows.append(build_error_row(approval, str(exc)))
            continue

        mapping_errors: list[str] = []
        if account_book_mapping and approval.account_book_name not in account_book_mapping:
            mapping_errors.append(f"账簿名称未映射: {approval.account_book_name}")
        if customer_mapping and approval.customer and approval.customer not in customer_mapping:
            mapping_errors.append(f"客户名称未映射: {approval.customer}")
        if mapping_errors:
            error_rows.append(build_error_row(approval, "; ".join(mapping_errors)))

        bank_account: str | None = None
        bank_match_error: str | None = None
        try:
            approval_time = base.parse_last_approval_time(approval.approval_record)
        except ValueError:
            approval_time = approval.salary_period_start

        for entry in entries:
            if entry.account_code == "1002" and entry.credit != Decimal("0"):
                if bank_mapping:
                    bank_account = base.find_bank_account_by_credit(entry.credit, approval_time, bank_mapping)
                    if not bank_account:
                        bank_match_error = f"银行账号未匹配: 贷方金额={entry.credit}, 日期={approval_time.strftime('%Y-%m-%d')}"
                break
        if bank_match_error:
            error_rows.append(build_error_row(approval, bank_match_error))

        for index, entry in enumerate(entries):
            output_rows.append(
                base.build_voucher_row(
                    template_row=template_row,
                    header_index=header_index,
                    approval=approval,
                    entry=entry,
                    voucher_seq=base_voucher_seq + voucher_offset,
                    voucher_no=base_voucher_no + voucher_offset,
                    entry_seq=base_entry_seq + entry_offset,
                    is_first_entry=index == 0,
                    account_book_mapping=account_book_mapping,
                    customer_mapping=customer_mapping,
                    bank_account=bank_account,
                )
            )
            entry_offset += 1
        voucher_offset += 1

    for row in output_rows:
        voucher_ws.append(row)

    amount_cols = []
    for col_idx in range(1, voucher_ws.max_column + 1):
        header = voucher_ws.cell(row=1, column=col_idx).value
        if header and ("FAMOUNTFOR" in str(header) or "FDEBIT" in str(header) or "FCREDIT" in str(header)):
            amount_cols.append(col_idx)
    for col_idx in amount_cols:
        for row_idx in range(3, voucher_ws.max_row + 1):
            cell = voucher_ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    if error_rows:
        error_ws = voucher_wb.create_sheet(title="错误明细")
        error_ws.append(ERROR_HEADERS)
        for row in error_rows:
            error_ws.append(row)
        for row_idx in range(2, error_ws.max_row + 1):
            for col_idx in [6, 7]:
                cell = error_ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

    output_dir.mkdir(parents=True, exist_ok=True)
    voucher_output_path = output_dir / "工资发放管理帐凭证.xlsx"
    voucher_wb.save(voucher_output_path)

    return {
        "approval_count": voucher_offset,
        "entry_count": len(output_rows),
        "error_count": len(error_rows),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成工资发放管理帐凭证表")
    parser.add_argument("--approval-path", default="审批表.xlsx")
    parser.add_argument("--template-path", default="凭证表.xlsx")
    parser.add_argument("--output-dir", default="生成结果")
    parser.add_argument("--account-book-mapping-path", default=ACCOUNT_BOOK_MAPPING_FILE)
    parser.add_argument("--customer-mapping-path", default=CUSTOMER_MAPPING_FILE)
    parser.add_argument("--target-profit-center", default=TARGET_PROFIT_CENTER)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = generate_salary_payment_management_voucher_files(
        approval_path=Path(args.approval_path),
        template_path=Path(args.template_path),
        output_dir=Path(args.output_dir),
        account_book_mapping_path=Path(args.account_book_mapping_path),
        customer_mapping_path=Path(args.customer_mapping_path),
        target_profit_center=args.target_profit_center,
    )
    print(f"目标归属利润中心: {args.target_profit_center}")
    print(f"成功生成审批单数量: {result['approval_count']}")
    print(f"生成凭证分录数量: {result['entry_count']}")
    print(f"错误数量: {result['error_count']}")
    print("输出路径: 生成结果/工资发放管理帐凭证.xlsx")
    print("  - 凭证sheet: 凭证#单据头(FBillHead)")
    print("  - 错误sheet: 错误明细")


if __name__ == "__main__":
    main()
