from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from copy import deepcopy
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers


@dataclass(frozen=True)
class ApprovalRow:
    approval_no: str
    customer: str
    contract_business_type: str
    account_book_name: str
    salary_period_start: datetime
    net_salary: Decimal
    tax: Decimal
    approval_record: str
    gross_salary: Decimal = Decimal("0")  # 应发工资
    social_insurance_personal: Decimal = Decimal("0")  # 社保和公积金个人部分
    after_tax_deduction: Decimal = Decimal("0")  # 税后扣除数
    amount_check: Decimal = Decimal("0")  # 金额检验（必须为0）


@dataclass(frozen=True)
class EntryRow:
    account_code: str
    debit: Decimal
    credit: Decimal
    customer_name: str | None
    is_tax: bool = False  # 标记是否为个税相关分录


SUPPORTED_CONTRACT_ACCOUNT_CODES = {
    "承揽业务": "6401.04.01",
    "派遣业务": "6401.04.01",
    "代理业务": "6401.02.01",
    # 助残业务使用特殊科目编码，在build_entry_rows中单独处理
}

# 科目编码到名称的映射
ACCOUNT_CODE_NAME_MAP = {
    "1002": "银行存款",
    "2221.13": "应交税费_代扣个人所得税",
    "6401.04.01": "主营业务成本_承揽成本_工资/经营所得",
    "6401.02.01": "主营业务成本_代理成本_工资",
    # 助残业务科目编码
    "6602.01.01.01": "管理费用_人力成本_工资薪酬_工资",
    "2211.01.01": "应付职工薪酬_工资薪金_工资",
    "2211.02.02": "应付职工薪酬_社保_社保个人部分",
}

APPROVAL_SHEET_NAME = "工资报酬支付申请"
VOUCHER_SHEET_NAME = "凭证#单据头(FBillHead)"
ERROR_HEADERS = ["审批编号", "客户", "合同业务类型", "工资所属期起", "实发工资", "个税", "错误原因"]

# 映射表文件名
ACCOUNT_BOOK_MAPPING_FILE = "公司主体账套号.xlsx"
CUSTOMER_MAPPING_FILE = "客户账套.xlsx"
BANK_JOURNAL_FILE = "银行存款日记账.xlsx"


def _normalize_bank_journal_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def _get_bank_journal_row_value(row: tuple[object, ...], index: int) -> object:
    return row[index] if index < len(row) else None


def _parse_bank_journal_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    raise ValueError(f"银行日记账日期解析失败: {value}")


def _locate_bank_journal_columns(ws) -> tuple[int, dict[str, int]]:
    required_headers = ("银行账号", "贷方金额", "业务日期")
    normalized_required = {name: _normalize_bank_journal_header(name) for name in required_headers}

    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True),
        start=1,
    ):
        header_index: dict[str, int] = {}
        for column_index, cell_value in enumerate(row):
            normalized_cell = _normalize_bank_journal_header(cell_value)
            for header_name, normalized_header in normalized_required.items():
                if normalized_cell == normalized_header:
                    header_index[header_name] = column_index
        if len(header_index) == len(required_headers):
            return row_index, header_index

    raise ValueError("银行存款日记账缺少必要表头: 银行账号、贷方金额、业务日期")


def load_bank_account_mapping(journal_path: Path) -> dict[tuple[Decimal, datetime.date], str]:
    """加载银行存款日记账，建立贷方金额+业务日期到银行账号的映射"""
    wb = load_workbook(journal_path, data_only=True)
    ws = wb.active

    mapping: dict[tuple[Decimal, datetime.date], str] = {}
    header_row, header_index = _locate_bank_journal_columns(ws)

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue

        bank_account = _get_bank_journal_row_value(row, header_index["银行账号"])
        credit_amount = _get_bank_journal_row_value(row, header_index["贷方金额"])
        business_date = _get_bank_journal_row_value(row, header_index["业务日期"])
        if bank_account and credit_amount is not None and business_date:
            try:
                amount = Decimal(str(credit_amount))
                date_val = _parse_bank_journal_date(business_date)
                mapping[(amount, date_val)] = str(bank_account)
            except (InvalidOperation, ValueError):
                continue

    return mapping


def find_bank_account_by_credit(
    credit_amount: Decimal,
    approval_date: datetime,
    bank_mapping: dict[tuple[Decimal, datetime.date], str],
) -> str | None:
    """根据贷方金额和审批日期查找银行账号"""
    # 先尝试精确匹配（金额+日期）
    approval_date_val = approval_date.date() if isinstance(approval_date, datetime) else approval_date

    # 规范化金额到2位小数，忽略小数点位数差异（如187364.6和187364.60视为相同）
    normalized_amount = round(credit_amount, 2)

    # 查找匹配（使用规范化后的金额比较）
    for (amt, date), acc in bank_mapping.items():
        normalized_amt = round(amt, 2)
        if normalized_amt == normalized_amount and date == approval_date_val:
            return acc

    # 如果日期匹配不到，只按金额查找所有可能的银行账号
    matches = [(amt, date, acc) for (amt, date), acc in bank_mapping.items() if round(amt, 2) == normalized_amount]
    if len(matches) == 1:
        return matches[0][2]
    elif len(matches) > 1:
        # 多个匹配，尝试找最接近日期的
        for amt, date, acc in matches:
            if date == approval_date_val:
                return acc
        # 没找到同日期的，返回第一个（但这种情况应该报错）
        return None

    return None


def load_account_book_mapping(mapping_path: Path) -> dict[str, tuple[str, str]]:
    """加载账簿名称到单据头序号和账簿编码的映射"""
    wb = load_workbook(mapping_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers) if name}

    # FName列是账簿名称，FNumber列是账簿编码，第一列是单据头序号
    mapping: dict[str, tuple[str, str]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):  # 第2行是中文描述，第3行开始是数据
        if not any(row):
            continue
        book_name = row[header_index.get("FName", 16)] if "FName" in header_index else row[16]
        book_seq = row[0]  # FBillHead(BD_AccountBook) 在第一列
        book_code = row[header_index.get("FNumber", 6)] if "FNumber" in header_index else row[6]
        if book_name:
            mapping[str(book_name).strip()] = (str(book_seq), str(book_code) if book_code else "")
    return mapping


def load_customer_mapping(mapping_path: Path) -> dict[str, str]:
    """加载客户名称到客户编码的映射"""
    wb = load_workbook(mapping_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers) if name}

    mapping: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        customer_code = row[header_index.get("客户编码", 0)]
        customer_name = row[header_index.get("客户名称", 1)]
        if customer_name and customer_code:
            mapping[str(customer_name).strip()] = str(customer_code).strip()
    return mapping


def parse_money(value: object) -> Decimal:
    """解析金额，去除货币符号和千分位分隔符"""
    text = "" if value is None else str(value).strip()
    cleaned = re.sub(r"[^0-9.\-]", "", text)
    if not cleaned:
        return Decimal("0")
    try:
        return Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError(f"金额解析失败: {value}") from exc


def parse_last_approval_time(record: str) -> datetime:
    """从审批记录中提取最后一条审批时间"""
    # 处理可能包含换行符的情况
    record = record.replace("\n", "")
    for chunk in reversed([item.strip() for item in record.split(";") if item.strip()]):
        parts = [part.strip() for part in chunk.split("|")]
        if len(parts) >= 3 and parts[2]:
            try:
                return datetime.strptime(parts[2], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
    raise ValueError("审批记录解析失败")


def parse_datetime_text(value: object) -> datetime:
    """解析日期时间文本"""
    text = str(value).strip() if value else ""
    if not text:
        raise ValueError("日期时间为空")
    # 尝试不同格式
    formats = ["%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d 00:00"]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"日期时间解析失败: {value}")


def build_explanation(approval: ApprovalRow) -> str:
    """生成摘要文本：付客户X月工资"""
    return (
        f"付{approval.customer}"
        f"{approval.salary_period_start.month}月工资"
    )


def resolve_expense_account_code(contract_business_type: str) -> str:
    """根据合同业务类型获取费用科目编码"""
    try:
        return SUPPORTED_CONTRACT_ACCOUNT_CODES[contract_business_type]
    except KeyError as exc:
        raise ValueError(f"不支持的合同业务类型: {contract_business_type}") from exc


def build_entry_rows(approval: ApprovalRow) -> list[EntryRow]:
    """将审批记录展开为凭证分录"""
    # 助残业务使用特殊科目编码规则
    if approval.contract_business_type == "助残业务":
        # 计算：应发工资 - 税后扣除数
        management_fee = approval.gross_salary - approval.after_tax_deduction
        entries = [
            # 1. 管理费用_人力成本_工资薪酬_工资 - 借方 = 应发工资 - 税后扣除数（不填客户名）
            EntryRow("6602.01.01.01", management_fee, Decimal("0"), None),
            # 2. 应付职工薪酬_工资薪金_工资 - 借方 = 实发工资（不填客户名）
            EntryRow("2211.01.01", approval.net_salary, Decimal("0"), None),
            # 3. 应付职工薪酬_社保_社保个人部分 - 贷方 = 社保和公积金个人部分（不填客户名）
            EntryRow("2211.02.02", Decimal("0"), approval.social_insurance_personal, None),
            # 4. 应付职工薪酬_工资薪金_工资 - 贷方 = 实发工资（不填客户名）
            EntryRow("2211.01.01", Decimal("0"), approval.net_salary, None),
            # 5. 银行存款 - 贷方 = 实发工资（银行存款需要填写客户名）
            EntryRow("1002", Decimal("0"), approval.net_salary, approval.customer),
        ]
        return entries

    # 其他业务类型的处理
    expense_code = resolve_expense_account_code(approval.contract_business_type)
    entries = [
        EntryRow(expense_code, approval.net_salary, Decimal("0"), approval.customer),
        EntryRow("1002", Decimal("0"), approval.net_salary, approval.customer),
    ]
    if approval.tax != Decimal("0"):
        entries.extend(
            [
                EntryRow(expense_code, approval.tax, Decimal("0"), approval.customer, is_tax=True),
                EntryRow("2221.13", Decimal("0"), approval.tax, None, is_tax=True),  # 2221.13不需要客户维度
            ]
        )
    return entries


def load_approval_rows(approval_path: Path) -> tuple[list[ApprovalRow], list[list[object]]]:
    """读取审批表数据"""
    wb = load_workbook(approval_path, data_only=True)
    ws = wb[APPROVAL_SHEET_NAME]
    headers = [cell.value for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers) if name}
    approvals: list[ApprovalRow] = []
    errors: list[list[object]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        approval_no_val = row[header_index.get("审批编号", 0)]
        if not approval_no_val:
            continue

        try:
            approvals.append(
                ApprovalRow(
                    approval_no=str(approval_no_val).strip(),
                    customer=str(row[header_index.get("客户", 5)] or "").strip(),
                    contract_business_type=str(row[header_index.get("合同业务类型", 7)] or "").strip(),
                    account_book_name=str(row[header_index.get("发放和申报主体", 9)] or "").strip(),
                    salary_period_start=parse_datetime_text(row[header_index.get("工资所属期起", 14)]),
                    net_salary=parse_money(row[header_index.get("实发工资", 24)]),
                    tax=parse_money(row[header_index.get("个税", 22)]),
                    approval_record=str(row[header_index.get("审批记录", 53)] or "").strip(),
                    gross_salary=parse_money(row[header_index.get("应发工资", 20)]),
                    social_insurance_personal=parse_money(row[header_index.get("社保和公积金个人部分", 21)]),
                    after_tax_deduction=parse_money(row[header_index.get("税后扣除数", 23)]),
                    amount_check=parse_money(row[header_index.get("金额检验（必须为0）", 25)]),
                )
            )
        except ValueError as exc:
            errors.append(
                [
                    row[header_index.get("审批编号", 0)],
                    row[header_index.get("客户", 5)],
                    row[header_index.get("合同业务类型", 7)],
                    row[header_index.get("工资所属期起", 14)],
                    row[header_index.get("实发工资", 24)],
                    row[header_index.get("个税", 22)],
                    str(exc),
                ]
            )
    return approvals, errors


def build_error_row(approval: ApprovalRow, reason: str) -> list[object]:
    """构建错误表行"""
    return [
        approval.approval_no,
        approval.customer,
        approval.contract_business_type,
        approval.salary_period_start.strftime("%Y-%m-%d %H:%M"),
        round(float(approval.net_salary), 2),
        round(float(approval.tax), 2),
        reason,
    ]


def build_voucher_row(
    template_row: list[object],
    header_index: dict[str, int],
    approval: ApprovalRow,
    entry: EntryRow,
    voucher_seq: int,
    voucher_no: int,
    entry_seq: int,
    is_first_entry: bool,
    account_book_mapping: dict[str, tuple[str, str]] | None = None,
    customer_mapping: dict[str, str] | None = None,
    bank_account: str | None = None,
) -> list[object]:
    """构建凭证行"""
    row = list(template_row)  # 复制模板行

    # 非第一条分录清空单据头字段
    if not is_first_entry:
        for key in [
            "FBillHead(GL_VOUCHER)", "FAccountBookID", "FAccountBookID#Name", "FDate", "FBUSDATE",
            "FYEAR", "FPERIOD", "FVOUCHERGROUPID", "FVOUCHERGROUPID#Name", "FVOUCHERGROUPNO",
            "FATTACHMENTS", "FISADJUSTVOUCHER", "FACCBOOKORGID", "FACCBOOKORGID#Name",
            "FSourceBillKey", "FSourceBillKey#Name", "FIMPORTVERSION", "F_NGU_CKXX", "*Split*1",
        ]:
            if key in header_index:
                row[header_index[key]] = None

    # 解析审批时间
    try:
        approval_time = parse_last_approval_time(approval.approval_record)
    except ValueError:
        # 如果审批记录解析失败，使用工资所属期起作为默认日期
        approval_time = approval.salary_period_start

    # 写入单据头字段（仅第一条分录）
    if is_first_entry:
        # 根据账簿名称映射获取单据头序号和账簿编码
        book_seq, book_code = ("", "")
        if account_book_mapping and approval.account_book_name in account_book_mapping:
            book_seq, book_code = account_book_mapping[approval.account_book_name]

        if "FBillHead(GL_VOUCHER)" in header_index:
            row[header_index["FBillHead(GL_VOUCHER)"]] = book_seq if book_seq else None
        if "FAccountBookID" in header_index:
            row[header_index["FAccountBookID"]] = book_code if book_code else None
        if "FAccountBookID#Name" in header_index:
            row[header_index["FAccountBookID#Name"]] = approval.account_book_name
        if "FDate" in header_index:
            row[header_index["FDate"]] = approval_time.strftime("%Y-%m-%d")
        if "FBUSDATE" in header_index:
            row[header_index["FBUSDATE"]] = approval_time.strftime("%Y-%m-%d")
        # 从日期获取会计年度和期间
        if "FYEAR" in header_index:
            row[header_index["FYEAR"]] = approval_time.year
        if "FPERIOD" in header_index:
            row[header_index["FPERIOD"]] = approval_time.month
        if "FVOUCHERGROUPNO" in header_index:
            row[header_index["FVOUCHERGROUPNO"]] = voucher_no

    # 写入分录字段
    if "FEntity" in header_index:
        row[header_index["FEntity"]] = None  # 清空分录序号
    if "FEXPLANATION" in header_index:
        explanation = build_explanation(approval)
        if entry.is_tax:
            explanation += "个税"
        row[header_index["FEXPLANATION"]] = explanation
    if "FACCOUNTID" in header_index:
        row[header_index["FACCOUNTID"]] = entry.account_code
    if "FACCOUNTID#Name" in header_index:
        # 填入科目编码对应的名称
        row[header_index["FACCOUNTID#Name"]] = ACCOUNT_CODE_NAME_MAP.get(entry.account_code)
    # 根据客户名称映射获取客户编码（如果没有客户名称或找不到映射，则为空）
    if "FDetailID#FFlex6" in header_index:
        customer_code = None
        if entry.customer_name and customer_mapping and entry.customer_name in customer_mapping:
            customer_code = customer_mapping[entry.customer_name]
        row[header_index["FDetailID#FFlex6"]] = customer_code
    if "FDetailID#FFlex6#Name" in header_index:
        row[header_index["FDetailID#FFlex6#Name"]] = entry.customer_name
    # 金额保留两位小数
    if "FAMOUNTFOR" in header_index:
        amount_val = round(float(entry.debit or entry.credit), 2)
        row[header_index["FAMOUNTFOR"]] = amount_val if amount_val != 0 else None
    if "FDEBIT" in header_index:
        debit_val = round(float(entry.debit), 2) if entry.debit != Decimal("0") else None
        row[header_index["FDEBIT"]] = debit_val
    if "FCREDIT" in header_index:
        credit_val = round(float(entry.credit), 2) if entry.credit != Decimal("0") else None
        row[header_index["FCREDIT"]] = credit_val
    if "FEXPORTENTRYID" in header_index:
        row[header_index["FEXPORTENTRYID"]] = None  # 清空现金流量分录ID
    # 银行存款科目时填入银行账号
    if entry.account_code == "1002" and bank_account:
        if "FDetailID#FF100006" in header_index:
            row[header_index["FDetailID#FF100006"]] = bank_account

    return row


def write_error_workbook(error_rows: list[list[object]], output_path: Path) -> None:
    """写入错误表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "错误明细"
    ws.append(ERROR_HEADERS)
    for row in error_rows:
        ws.append(row)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def generate_salary_voucher_files(
    approval_path: Path,
    template_path: Path,
    output_dir: Path,
    account_book_mapping_path: Path | None = None,
    customer_mapping_path: Path | None = None,
) -> dict[str, int]:
    """生成凭证文件和错误文件"""
    approvals, preload_errors = load_approval_rows(approval_path)

    # 加载映射表
    account_book_mapping: dict[str, tuple[str, str]] | None = None
    customer_mapping: dict[str, str] | None = None
    bank_mapping: dict[tuple[Decimal, datetime.date], str] | None = None

    if account_book_mapping_path and account_book_mapping_path.exists():
        account_book_mapping = load_account_book_mapping(account_book_mapping_path)
    if customer_mapping_path and customer_mapping_path.exists():
        customer_mapping = load_customer_mapping(customer_mapping_path)
    # 加载银行存款日记账映射
    bank_journal_path = Path(BANK_JOURNAL_FILE)
    if bank_journal_path.exists():
        bank_mapping = load_bank_account_mapping(bank_journal_path)

    voucher_wb = load_workbook(template_path)
    voucher_ws = voucher_wb[VOUCHER_SHEET_NAME]

    headers = [cell.value for cell in voucher_ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers) if name}

    template_row = [cell.value for cell in voucher_ws[3]]
    base_voucher_seq = int(template_row[header_index.get("FBillHead(GL_VOUCHER)", 0)] or 0)
    base_voucher_no = int(template_row[header_index.get("FVOUCHERGROUPNO", 9)] or 0)
    base_entry_seq = int(template_row[header_index.get("FEntity", 19)] or 0)

    # 删除模板原有数据行（保留表头）
    voucher_ws.delete_rows(3, voucher_ws.max_row - 2)

    output_rows: list[list[object]] = []
    error_rows = list(preload_errors)
    voucher_offset = 0
    entry_offset = 0

    for approval in approvals:
        # 检查金额检验（必须为0）
        if approval.amount_check != Decimal("0"):
            error_rows.append(build_error_row(approval, f"金额检验不为0: {approval.amount_check}"))
            continue

        try:
            entries = build_entry_rows(approval)
        except ValueError as exc:
            error_rows.append(build_error_row(approval, str(exc)))
            continue

        # 检查映射情况
        mapping_errors: list[str] = []

        # 检查账簿映射
        if account_book_mapping and approval.account_book_name not in account_book_mapping:
            mapping_errors.append(f"账簿名称未映射: {approval.account_book_name}")

        # 检查客户映射（对于有客户名称的审批）
        if customer_mapping and approval.customer and approval.customer not in customer_mapping:
            mapping_errors.append(f"客户名称未映射: {approval.customer}")

        # 如果有映射错误，记录到错误表
        if mapping_errors:
            error_rows.append(build_error_row(approval, "; ".join(mapping_errors)))

        # 查找银行账号（根据银行存款的贷方金额和审批日期）
        bank_account: str | None = None
        bank_match_error: str | None = None
        try:
            approval_time = parse_last_approval_time(approval.approval_record)
        except ValueError:
            approval_time = approval.salary_period_start

        # 找到银行存款科目（1002）的贷方金额
        for entry in entries:
            if entry.account_code == "1002" and entry.credit != Decimal("0"):
                if bank_mapping:
                    bank_account = find_bank_account_by_credit(entry.credit, approval_time, bank_mapping)
                    if not bank_account:
                        bank_match_error = f"银行账号未匹配: 贷方金额={entry.credit}, 日期={approval_time.strftime('%Y-%m-%d')}"
                break

        # 如果银行账号匹配失败，记录错误
        if bank_match_error:
            error_rows.append(build_error_row(approval, bank_match_error))

        for index, entry in enumerate(entries):
            output_rows.append(
                build_voucher_row(
                    template_row=template_row,
                    header_index=header_index,
                    approval=approval,
                    entry=entry,
                    voucher_seq=base_voucher_seq + voucher_offset,
                    voucher_no=base_voucher_no + voucher_offset,
                    entry_seq=base_entry_seq + entry_offset,
                    is_first_entry=index == 0,
                    account_book_mapping=account_book_mapping,
                    customer_mapping=customer_mapping,
                    bank_account=bank_account,
                )
            )
            entry_offset += 1
        voucher_offset += 1

    for row in output_rows:
        voucher_ws.append(row)

    # 设置金额列格式为显示两位小数
    # 查找金额列索引
    amount_cols = []
    for col_idx in range(1, voucher_ws.max_column + 1):
        header = voucher_ws.cell(row=1, column=col_idx).value
        if header and ('FAMOUNTFOR' in str(header) or 'FDEBIT' in str(header) or 'FCREDIT' in str(header)):
            amount_cols.append(col_idx)

    # 设置金额列格式
    for col_idx in amount_cols:
        for row_idx in range(3, voucher_ws.max_row + 1):
            cell = voucher_ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    # 在同一个Excel文件中添加错误sheet
    if error_rows:
        error_ws = voucher_wb.create_sheet(title="错误明细")
        error_ws.append(ERROR_HEADERS)
        for row in error_rows:
            error_ws.append(row)

        # 设置错误表中的金额列格式（实发工资列5、个税列6）
        for row_idx in range(2, error_ws.max_row + 1):
            for col_idx in [5, 6]:  # 实发工资和个税列
                cell = error_ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

    output_dir.mkdir(parents=True, exist_ok=True)
    voucher_output_path = output_dir / "工资报酬凭证.xlsx"

    voucher_wb.save(voucher_output_path)

    return {
        "approval_count": voucher_offset,
        "entry_count": len(output_rows),
        "error_count": len(error_rows),
    }


def main() -> None:
    """主函数入口"""
    result = generate_salary_voucher_files(
        approval_path=Path("审批表.xlsx"),
        template_path=Path("凭证表.xlsx"),
        output_dir=Path("生成结果"),
        account_book_mapping_path=Path(ACCOUNT_BOOK_MAPPING_FILE),
        customer_mapping_path=Path(CUSTOMER_MAPPING_FILE),
    )
    print(f"成功生成审批单数量: {result['approval_count']}")
    print(f"生成凭证分录数量: {result['entry_count']}")
    print(f"错误数量: {result['error_count']}")
    print("输出路径: 生成结果/工资报酬凭证.xlsx")
    print("  - 凭证sheet: 凭证#单据头(FBillHead)")
    print("  - 错误sheet: 错误明细")
    print("错误表输出路径: 生成结果/工资报酬凭证错误.xlsx")


if __name__ == "__main__":
    main()
